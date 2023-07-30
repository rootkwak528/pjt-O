import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

from src.rule import Rule


def find_header_row_from_sheet(sheet: str):
    for i in range(1, sheet.max_row + 1):
        if sheet[f'E{i}'].value is not None:
            return i


def process(rule_file_name: str, original_file_name: str, output_file_name: str):

    # pre_process
    rule_df = pd.read_excel(rule_file_name, sheet_name='rule', header=1)

    rules = [
        Rule(_row['구분1'], _row['구분2'], _row['구분3'], _row['적요조건'], _row['작성자조건'])
        for _, _row in rule_df.iterrows() if _row["제외코드"] not in ["예시", "제외"]
    ]

    # validation
    for rule in rules:
        if not rule.validate():
            return

    # process
    wb = load_workbook(filename=original_file_name)
    sheet = wb["비용현황"]

    header_row = find_header_row_from_sheet(sheet)
    header = {cell.value: idx for idx, cell in enumerate(sheet[header_row])}

    for row in tqdm(tuple(sheet.iter_rows(min_row=header_row+1))):
        for rule in rules:
            if rule.desc_rule(row[header['적요']].value) and rule.writer_rule(row[header['작성자']].value):
                row[header['구분1']].value = rule.div1
                row[header['구분2']].value = rule.div2
                row[header['구분3']].value = rule.div3
                break

    wb.save(output_file_name)
